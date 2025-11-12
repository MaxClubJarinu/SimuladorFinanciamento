# app_financiamento.py
import streamlit as st
st.set_page_config(page_title="Simulador de financiamento imobiliário", layout="centered")


# === Logo no canto (com caminho relativo ao arquivo) ===
# === Logo no canto (com caminho relativo ao arquivo + offsets) ===
from base64 import b64encode
from pathlib import Path

ASSETS_DIR = Path(__file__).parent / "assets"
LOGO_PATH = ASSETS_DIR / "logo.jpg"
# === assets auxiliares do login ===
HERO_PATH = ASSETS_DIR / "login_bg.jpg"  # sua imagem centralizada

def _b64(path: Path) -> str:
    return b64encode(path.read_bytes()).decode()

def add_corner_image(image_path: Path, width_px: int = 90, corner: str = "top-left",
                     offset_x: int = 60, offset_y: int = 50):
    """Exibe uma imagem fixa em um canto da página sem mexer no layout."""
    try:
        data = image_path.read_bytes()
        b64 = b64encode(data).decode()

        if corner == "top-right":
            pos_rules = f"top: {offset_y}px; right: {offset_x}px;"
            css_class = "corner-logo-tr"
        elif corner == "top-left":
            pos_rules = f"top: {offset_y}px; left: {offset_x}px;"
            css_class = "corner-logo-tl"
        elif corner == "bottom-right":
            pos_rules = f"bottom: {offset_y}px; right: {offset_x}px;"
            css_class = "corner-logo-br"
        else:
            pos_rules = f"bottom: {offset_y}px; left: {offset_x}px;"
            css_class = "corner-logo-bl"

        st.markdown(
            f"""
            <style>
            .{css_class} {{
                position: fixed;
                {pos_rules}
                width: {width_px}px;
                z-index: 9999;
                border-radius: 8px;
                box-shadow: 0 2px 12px rgba(0,0,0,.15);
            }}
            @media (max-width: 768px) {{
                .{css_class} {{ width: {width_px}px; }}
            }}
            </style>
            <img class="{css_class}" src="data:image/jpeg;base64,{b64}">
            """,
            unsafe_allow_html=True
        )
    except FileNotFoundError:
        st.warning(f"Logo não encontrado em: {image_path}. Confirme se o arquivo existe no repositório.")
    except Exception as e:
        st.warning(f"Falha ao exibir o logo: {e}")



from pathlib import Path
from io import BytesIO
import calendar
from datetime import datetime as dt, time
from dateutil.relativedelta import relativedelta
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

# ==========================
# Utilidades e salvaguardas
# ==========================
HEADER_FILL = PatternFill(start_color="FFD3D3D3", end_color="FFD3D3D3", fill_type="solid")
DATE_FORMAT = 'dd/mm/yyyy'
CURRENCY_FORMAT = '"R$" #,##0.00'
PERCENT_FORMAT = '0.00%'

GREEN_COLOR = "FF00B050"  # verde
RED_COLOR   = "FFFF0000"  # vermelho

def load_taxas(filepath: str) -> dict:
    taxas = {}
    path = Path(filepath)
    if not path.exists():
        return taxas
    try:
        content = path.read_text(encoding='utf-8').strip()
    except Exception as e:
        st.error(f"Erro ao ler {filepath}: {e}")
        return taxas
    blocos = [b.strip() for b in content.split("\n\n") if b.strip()]
    for bloco in blocos:
        linhas = bloco.splitlines()
        if not linhas:
            continue
        nome = linhas[0].strip()
        taxas[nome] = {}
        for linha in linhas[1:]:
            if '=' in linha:
                chave, valor = linha.split('=', 1)
                chave = chave.strip()
                valor = valor.strip()
                try:
                    taxas[nome][chave] = float(valor)
                except ValueError:
                    taxas[nome][chave] = valor
    return taxas

def adjust_day(date, preferred_day):
    try:
        return date.replace(day=preferred_day)
    except ValueError:
        last = calendar.monthrange(date.year, date.month)[1]
        return date.replace(day=last)

def days_in_month(date):
    return calendar.monthrange(date.year, date.month)[1]

class PaymentTracker:
    def __init__(self, dia_pagamento, taxa_juros):
        self.last_date = None
        self.dia = dia_pagamento
        self.taxa = taxa_juros
    def calculate(self, current_date, saldo):
        if self.last_date is None:
            self.last_date = current_date
            return 0.0, 0, 0.0
        dias_corridos = (current_date - self.last_date).days
        taxa_efetiva = self.taxa * (dias_corridos / 30)
        juros = saldo * taxa_efetiva
        self.last_date = current_date
        return juros, dias_corridos, taxa_efetiva

def login_screen():
    # --- CSS para centralizar a imagem e sobrepor o título ---
    try:
        hero_b64 = _b64(HERO_PATH)
    except Exception:
        hero_b64 = ""  # se faltar a imagem, só mostra o título comum

    st.markdown(f"""
    <style>
      /* container geral do “hero” */
      .hero-wrap {{
        position: relative;
        width: min(650px, 95vw);     /* largura máxima da imagem */
        margin: 32px auto 16px;      /* centraliza horizontalmente */
        border-radius: 16px;
        overflow: hidden;
        box-shadow: 0 10px 28px rgba(0,0,0,.12);
      }}
      .hero-img {{
        width: 100%;
        display: block;
        height: auto;
      }}
      /* título centralizado sobre a imagem */
      .hero-title {{
        position: absolute;
        inset: 0;                    /* ocupa toda a área do hero */
        display: flex;
        align-items: center;
        justify-content: center;
        text-align: center;
        color: #ffffff;
        text-shadow: 0 2px 16px rgba(0,0,0,.45);
        font-weight: 300;
        font-size: clamp(28px, 4.2vw, 46px);
        padding: 8px;
      }}
      /* espaçamento do bloco do formulário */
      .login-form {{
        width: min(520px, 95vw);
        margin: 12px auto 24px;
        padding: 16px 0;
      }}
    </style>

    <div class="hero-wrap">
      <img class="hero-img" src="data:image/jpeg;base64,{hero_b64}">
      <!-- APAGUE a linha abaixo para ficar só a imagem -->
    </div>
    """, unsafe_allow_html=True)

    
    # --- Formulário de login (separado da imagem) ---
    with st.form(key="__login__", clear_on_submit=False):
        st.markdown('<div class="login-form">', unsafe_allow_html=True)
        user = st.text_input("Login", placeholder="Digite seu login")
        pwd  = st.text_input("Senha", type="password", placeholder="Digite sua senha")
        ok   = st.form_submit_button("Entrar", use_container_width=True)
        st.markdown('</div>', unsafe_allow_html=True)


    # ===== Lógica do login (mesma de antes) =====
    if ok:
        if user == "Max Club Jarinu" and pwd == "maxclub123":
            st.session_state.authenticated = True
            st.rerun()
        else:
            st.error("Login ou senha incorretos. Tente novamente.")


# ==========================
# Corpo do App (em abas)
# ==========================
def app_body():
    st.title("Simulador de financiamento imobiliário")
    st.caption("Preencha os dados nas abas abaixo e clique em **Gerar Planilha** para baixar o Excel.")

    # Carrega taxas com fallback
    taxas_por_emp = load_taxas('taxas.txt')
    if not taxas_por_emp:
        taxas_por_emp = {
            "Padrão (fallback)": {
                "TAXA_EMISSAO_CCB": 500.0,
                "TAXA_EMISSAO_CONTRATO_ALIENACAO_FIDUCIARIA": 350.0,
                "TAXA_REGISTRO_IMOVEL": 1200.0,
                "TAXA_ESCRITURA_IMOVEL": 900.0,
                "TAXA_SEGURO_PRESTAMISTA_PCT": 0.012,
                "TAXA_INCC": 0.006,
                "TAXA_IPCA": 0.004,
                "taxa_pre": 0.02,
                "taxa_pos": 0.018
            }
        }
        st.info(
            "Arquivo **taxas.txt** não encontrado. Usando taxas padrão (fallback) para simular. "
            "Depois, coloque seu `taxas.txt` na mesma pasta para utilizar suas taxas reais."
        )

    tab1, tab2, tab3 = st.tabs(["1) Dados do contrato", "2) Pagamentos extras", "3) Gerar planilha"])

    # ====== Aba 1: Dados do contrato ======
    with tab1:
        st.subheader("Informações básicas")
        col1, col2 = st.columns([2, 1])
        cliente = col1.text_input("Nome do cliente", placeholder="Ex.: João da Silva", help="Identificação que aparecerá na planilha.")
        dia_pagamento = col2.number_input("Dia da parcela (1-31)", min_value=1, max_value=31, value=10, step=1,
                                          help="Dia do mês em que o cliente prefere pagar as parcelas.")

        col3, col4 = st.columns([1, 2])
        valor_imovel = col3.number_input("Valor total do imóvel (R$)", min_value=0.0, step=0.01, format="%.2f", value=0.0,
                                         help="Preço total do imóvel.")
        empreendimento = col4.selectbox("Empreendimento", options=list(taxas_por_emp.keys()),
                                        help="Selecione o empreendimento/projeto.")

        # Extrai taxas do empreendimento escolhido
        taxas_sel = taxas_por_emp.get(empreendimento, {})
        TAXA_EMISSAO_CCB = taxas_sel.get('TAXA_EMISSAO_CCB', 0.0)
        TAXA_EMISSAO_CONTRATO_ALIENACAO_FIDUCIARIA = taxas_sel.get('TAXA_EMISSAO_CONTRATO_ALIENACAO_FIDUCIARIA', 0.0)
        TAXA_REGISTRO_IMOVEL = taxas_sel.get('TAXA_REGISTRO_IMOVEL', 0.0)
        TAXA_ESCRITURA_IMOVEL = taxas_sel.get('TAXA_ESCRITURA_IMOVEL', 0.0)
        TAXA_SEGURO_PRESTAMISTA_PCT = taxas_sel.get('TAXA_SEGURO_PRESTAMISTA_PCT', 0.0)
        TAXA_INCC = taxas_sel.get('TAXA_INCC', 0.0)
        TAXA_IPCA = taxas_sel.get('TAXA_IPCA', 0.0)
        taxa_pre = taxas_sel.get('taxa_pre', 0.0)
        taxa_pos = taxas_sel.get('taxa_pos', 0.0)

        # Extras percentuais
        taxas_extras = []
        for chave, val in taxas_sel.items():
            if chave.endswith('_PCT') and chave not in ['TAXA_SEGURO_PRESTAMISTA_PCT']:
                periodo = 'pré-entrega da chave' if 'INCC' in chave else 'pós-entrega da chave'
                taxas_extras.append({'pct': val, 'periodo': periodo})

        st.markdown("### Datas-chave")
        cold1, cold2 = st.columns(2)
        data_base_date = cold1.date_input("Data da assinautra do contrato", value=dt.now().date())
        data_base = dt.combine(data_base_date, time())
        data_inicio_pre_date = cold2.date_input("Início dos pagamentos (antes da entrega das chaves)", value=dt.now().date())
        data_inicio_pre = dt.combine(data_inicio_pre_date, time())

        coldd1, coldd2 = st.columns(2)
        data_entrega_date = coldd1.date_input("Conclusão da obra e entrega das chaves", value=dt.now().date())
        data_entrega = dt.combine(data_entrega_date, time())

        st.markdown("### Capacidades de pagamento")
        colp1, colp2 = st.columns(2)
        capacidade_pre = colp1.number_input("Valor da parcela ANTES da conclusão da obra (R$)", min_value=0.0, step=0.01, value=0.0,
                                            help="Quanto o cliente consegue pagar por mês durante a obra.")
        capacidade_pos_antes = colp2.number_input("Valor da parcela DEPOIS da conclusão da obra (R$)", min_value=0.0, step=0.01, value=0.0,
                                                  help="Quanto o cliente consegue pagar por mês após a entrega das chaves.")

        cola, colb, colc = st.columns(3)
        fgts = cola.number_input("FGTS para abatimento (R$)", min_value=0.0, step=0.01, value=0.0)
        fin_banco = colb.number_input("Valor financiado pelo banco (R$)", min_value=0.0, step=0.01, value=0.0)
        val_parcela_banco = colc.number_input("Parcela mensal do banco (R$)", min_value=0.0, step=0.01, value=0.0)

        capacidade_pos = capacidade_pos_antes - val_parcela_banco
        if capacidade_pos_antes or val_parcela_banco:
            if capacidade_pos < 0:
                st.warning(f"Parcela do banco maior que a capacidade informada. Capacidade pós-chaves restante: R${capacidade_pos:.2f}.")
            else:
                st.info(f"Capacidade pós-chaves disponível (para a construtora): **R${capacidade_pos:.2f}**/mês.")

        st.session_state.inputs = {
            "cliente": cliente,
            "dia_pagamento": dia_pagamento,
            "valor_imovel": valor_imovel,
            "empreendimento": empreendimento,
            "taxas_sel": taxas_sel,
            "TAXA_EMISSAO_CCB": TAXA_EMISSAO_CCB,
            "TAXA_EMISSAO_CONTRATO_ALIENACAO_FIDUCIARIA": TAXA_EMISSAO_CONTRATO_ALIENACAO_FIDUCIARIA,
            "TAXA_REGISTRO_IMOVEL": TAXA_REGISTRO_IMOVEL,
            "TAXA_ESCRITURA_IMOVEL": TAXA_ESCRITURA_IMOVEL,
            "TAXA_SEGURO_PRESTAMISTA_PCT": TAXA_SEGURO_PRESTAMISTA_PCT,
            "TAXA_INCC": TAXA_INCC,
            "TAXA_IPCA": TAXA_IPCA,
            "taxa_pre": taxa_pre,
            "taxa_pos": taxa_pos,
            "taxas_extras": taxas_extras,
            "data_base": data_base,
            "data_inicio_pre": data_inicio_pre,
            "data_entrega": data_entrega,
            "capacidade_pre": capacidade_pre,
            "capacidade_pos_antes": capacidade_pos_antes,
            "val_parcela_banco": val_parcela_banco,
            "capacidade_pos": capacidade_pos,
            "fgts": fgts,
            "fin_banco": fin_banco
        }

    # ====== Aba 2: Pagamentos extras ======
    with tab2:
        st.subheader("Pagamentos adicionais únicos")
        n_non_rec = st.number_input("Quantidade de pagamentos adicionais únicos", min_value=0, step=1, value=0)
        non_rec = []
        for i in range(int(n_non_rec)):
            st.markdown(f"**Pagamento adicional {i+1}:**")
            col_nr1, col_nr2 = st.columns([1, 1])
            d_date = col_nr1.date_input("Data do pagamento", key=f"nr_d_{i}", value=dt.now().date())
            v = col_nr2.number_input("Valor (R$)", min_value=0.0, step=0.01, key=f"nr_v_{i}", value=0.0)
            col_nr3, col_nr4 = st.columns([2, 1])
            desc = col_nr3.text_input("Descrição", key=f"nr_desc_{i}", value=f"Pagamento adicional {i+1}")
            assoc = col_nr4.checkbox("Associar à parcela do mês?", key=f"nr_assoc_{i}", value=False,
                                     help="Se marcado, o pagamento cai no mesmo dia da parcela do mês.")
            d = dt.combine(d_date, time())
            if 'inputs' in st.session_state:
                d = adjust_day(d, st.session_state.inputs.get("dia_pagamento", 10)) if assoc else d
            non_rec.append({'data': d, 'tipo': desc, 'valor': v, 'assoc': assoc})

        st.subheader("Pagamentos semestrais")
        n_semi = st.number_input("Quantidade de séries semestrais", min_value=0, step=1, value=0)
        semi_series = []
        for i in range(int(n_semi)):
            st.markdown(f"**Série semestral {i+1}:**")
            col_s1, col_s2, col_s3 = st.columns([2, 1, 1])
            d0_date = col_s1.date_input("Data inicial", key=f"s_d0_{i}", value=dt.now().date())
            v = col_s2.number_input("Valor (R$)", min_value=0.0, step=0.01, key=f"s_v_{i}", value=0.0)
            assoc = col_s3.checkbox("Associar à parcela mensal?", key=f"s_assoc_{i}", value=False)
            semi_series.append({'d0': dt.combine(d0_date, time()), 'v': v, 'assoc': assoc, 'tipo': 'Pagamento Semestral'})

        st.subheader("Pagamentos anuais")
        n_ann = st.number_input("Quantidade de séries anuais", min_value=0, step=1, value=0)
        annual_series = []
        for i in range(int(n_ann)):
            st.markdown(f"**Série anual {i+1}:**")
            col_a1, col_a2, col_a3 = st.columns([2, 1, 1])
            d0_date = col_a1.date_input("Data inicial", key=f"a_d0_{i}", value=dt.now().date())
            v = col_a2.number_input("Valor (R$)", min_value=0.0, step=0.01, key=f"a_v_{i}", value=0.0)
            assoc = col_a3.checkbox("Associar à parcela mensal?", key=f"a_assoc_{i}", value=False)
            annual_series.append({'d0': dt.combine(d0_date, time()), 'v': v, 'assoc': assoc, 'tipo': 'Pagamento Anual'})

        st.session_state.extras = {
            "non_rec": non_rec,
            "semi_series": semi_series,
            "annual_series": annual_series
        }

    # ====== Aba 3: Gerar planilha ======
    with tab3:
        st.subheader("Geração da simulação (Excel)")
        st.write("Revise as abas anteriores. Quando estiver tudo certo, clique no botão abaixo.")

        if st.button("Gerar Planilha", type="primary"):
            try:
                if 'inputs' not in st.session_state:
                    st.error("Preencha a aba 'Dados do contrato' antes.")
                    st.stop()
                I = st.session_state.inputs
                E = st.session_state.get("extras", {"non_rec": [], "semi_series": [], "annual_series": []})

                cliente = I["cliente"]
                dia_pagamento = I["dia_pagamento"]
                valor_imovel = I["valor_imovel"]
                TAXA_EMISSAO_CCB = I["TAXA_EMISSAO_CCB"]
                TAXA_EMISSAO_CONTRATO_ALIENACAO_FIDUCIARIA = I["TAXA_EMISSAO_CONTRATO_ALIENACAO_FIDUCIARIA"]
                TAXA_REGISTRO_IMOVEL = I["TAXA_REGISTRO_IMOVEL"]
                TAXA_ESCRITURA_IMOVEL = I["TAXA_ESCRITURA_IMOVEL"]
                TAXA_SEGURO_PRESTAMISTA_PCT = I["TAXA_SEGURO_PRESTAMISTA_PCT"]
                TAXA_INCC = I["TAXA_INCC"]
                TAXA_IPCA = I["TAXA_IPCA"]
                taxa_pre = I["taxa_pre"]
                taxa_pos = I["taxa_pos"]
                taxas_extras = I["taxas_extras"]
                data_base = I["data_base"]
                data_inicio_pre = I["data_inicio_pre"]
                data_entrega = I["data_entrega"]
                capacidade_pre = I["capacidade_pre"]
                capacidade_pos_antes = I["capacidade_pos_antes"]
                val_parcela_banco = I["val_parcela_banco"]
                capacidade_pos = I["capacidade_pos"]
                fgts = I["fgts"]
                fin_banco = I["fin_banco"]

                # Copia listas de extras
                non_rec = list(E["non_rec"])
                semi_series = list(E["semi_series"])
                annual_series = list(E["annual_series"])

                # Agrega séries recorrentes (gera agenda bruta)
                for series in semi_series:
                    for n in range(100):
                        d = series['d0'] + relativedelta(months=6 * n)
                        if series['assoc']:
                            d = adjust_day(d, dia_pagamento)
                        non_rec.append({'data': d, 'tipo': 'Pagamento Semestral', 'valor': series['v'], 'assoc': series['assoc']})
                for series in annual_series:
                    for n in range(100):
                        d = series['d0'] + relativedelta(years=n)
                        if series['assoc']:
                            d = adjust_day(d, dia_pagamento)
                        non_rec.append({'data': d, 'tipo': 'Pagamento Anual', 'valor': series['v'], 'assoc': series['assoc']})

                # Separa pré/pós
                pre_nr = sorted([e for e in non_rec if e['data'] < data_entrega], key=lambda x: x['data'])
                post_nr = sorted([e for e in non_rec if e['data'] >= data_entrega], key=lambda x: x['data'])

                # Coletores
                eventos = []
                saldo = valor_imovel

                # Contagem de parcelas especiais (para exibir k/N)
                semi_total = len([e for e in non_rec if e['tipo'] == 'Pagamento Semestral'])
                annual_total = len([e for e in non_rec if e['tipo'] == 'Pagamento Anual'])
                semi_seq = 0
                annual_seq = 0

                # Data base
                eventos.append({
                    'data': data_base, 'parcela': '', 'tipo': 'Data-Base (assinatura do contrato)',
                    'valor': 0.0, 'juros': 0.0, 'dias_corridos': 0, 'taxa_efetiva': 0.0,
                    'incc': 0.0, 'ipca': 0.0, 'taxas_extra': [0.0] * len(taxas_extras),
                    'Total de mudança (R$)': 0.0, 'saldo': saldo
                })

                # ========== PRÉ-ENTREGA ==========
                tracker_pre = PaymentTracker(dia_pagamento, taxa_pre)
                tracker_pre.last_date = data_base

                prev_date = data_inicio_pre
                cursor = data_inicio_pre
                while True:
                    d_evt = adjust_day(cursor, dia_pagamento)
                    if d_evt >= data_entrega:
                        break

                    # (a) não-associados entre prev_date e d_evt — seguem iguais
                    for ev_nr in [e for e in pre_nr if not e['assoc'] and prev_date < e['data'] < d_evt]:
                        juros, dias_corr, taxa_eff = tracker_pre.calculate(ev_nr['data'], saldo)
                        incc_nr = saldo * TAXA_INCC
                        extras_nr = [saldo * t['pct'] if t['periodo'] in ['pré-entrega da chave', 'ambos'] else 0.0 for t in taxas_extras]
                        total_taxas_nr = sum(extras_nr) + incc_nr
                        abat_nr = ev_nr['valor'] - juros - total_taxas_nr
                        saldo -= abat_nr
                        eventos.append({
                            'data': ev_nr['data'], 'parcela': '', 'tipo': ev_nr['tipo'], 'valor': ev_nr['valor'],
                            'juros': juros, 'dias_corridos': dias_corr, 'taxa_efetiva': taxa_eff,
                            'incc': incc_nr, 'ipca': 0.0, 'taxas_extra': extras_nr,
                            'Total de mudança (R$)': -abat_nr, 'saldo': saldo
                        })

                    # >>> ORDEM CORRETA A PARTIR DAQUI <<<

                    # (b) parcela mensal pré — calcula encargos ANTES dos associados
                    juros, dias_corr, taxa_eff = tracker_pre.calculate(d_evt, saldo)
                    incc = saldo * TAXA_INCC
                    extras = [saldo * t['pct'] if t['periodo'] in ['pré-entrega da chave', 'ambos'] else 0.0 for t in taxas_extras]
                    total_taxas = sum(extras) + incc
                    valor_total = capacidade_pre
                    abat_mes = valor_total - juros - total_taxas
                    saldo -= abat_mes
                    eventos.append({
                        'data': d_evt, 'parcela': '', 'tipo': 'Pré-Entrega', 'valor': valor_total,
                        'juros': juros, 'dias_corridos': dias_corr, 'taxa_efetiva': taxa_eff,
                        'incc': incc, 'ipca': 0.0, 'taxas_extra': extras,
                        'Total de mudança (R$)': -abat_mes, 'saldo': saldo
                    })

                    # (c) associados do dia — linhas separadas, zerando encargos, DEPOIS da parcela
                    associados = [e for e in pre_nr if e['assoc'] and e['data'] == d_evt]
                    for ev_as in associados:
                        abat_assoc = ev_as['valor']  # 100% para principal
                        saldo -= abat_assoc
                        eventos.append({
                            'data': d_evt, 'parcela': '', 'tipo': ev_as['tipo'] + " (Associado)", 'valor': ev_as['valor'],
                            'juros': 0.0, 'dias_corridos': 0, 'taxa_efetiva': 0.0,
                            'incc': 0.0, 'ipca': 0.0, 'taxas_extra': [0.0]*len(taxas_extras),
                            'Total de mudança (R$)': -abat_assoc, 'saldo': saldo
                        })

                    prev_date = d_evt
                    cursor += relativedelta(months=1)


                # ========== ENTREGA ==========
                ent = data_entrega
                zero_extras = [0.0] * len(taxas_extras)

                for desc, v in [('Abatimento FGTS', fgts), ('Abatimento Fin. Banco', fin_banco)]:
                    saldo -= v
                    eventos.append({
                        'data': ent, 'parcela': '', 'tipo': desc, 'valor': 0.0,
                        'juros': 0.0, 'dias_corridos': '', 'taxa_efetiva': '',
                        'incc': 0.0, 'ipca': 0.0, 'taxas_extra': zero_extras,
                        'Total de mudança (R$)': -v,  # abatimento => negativo
                        'saldo': saldo
                    })

                for nome, val in [('Emissão CCB', TAXA_EMISSAO_CCB),
                                  ('Alienação Fiduciária', TAXA_EMISSAO_CONTRATO_ALIENACAO_FIDUCIARIA),
                                  ('Registro', TAXA_REGISTRO_IMOVEL),
                                  ('Escritura Imóvel', TAXA_ESCRITURA_IMOVEL)]:
                    saldo += val
                    eventos.append({
                        'data': ent, 'parcela': '', 'tipo': 'Taxa ' + nome, 'valor': 0.0,
                        'juros': 0.0, 'dias_corridos': '', 'taxa_efetiva': '',
                        'incc': 0.0, 'ipca': 0.0, 'taxas_extra': zero_extras,
                        'Total de mudança (R$)': val,  # adiciona saldo => positivo
                        'saldo': saldo
                    })

                fee = saldo * TAXA_SEGURO_PRESTAMISTA_PCT
                saldo += fee
                eventos.append({
                    'data': ent, 'parcela': '', 'tipo': 'Taxa Seguro Prestamista', 'valor': 0.0,
                    'juros': 0.0, 'dias_corridos': '', 'taxa_efetiva': '',
                    'incc': 0.0, 'ipca': 0.0, 'taxas_extra': zero_extras,
                    'Total de mudança (R$)': fee,  # adiciona saldo => positivo
                    'saldo': saldo
                })

                eventos.append({
                    'data': ent, 'parcela': '', 'tipo': 'Data da entrega das chaves', 'valor': 0.0,
                    'juros': 0.0, 'dias_corridos': 0, 'taxa_efetiva': 0.0,
                    'incc': 0.0, 'ipca': 0.0, 'taxas_extra': [0.0]*len(taxas_extras),
                    'Total de mudança (R$)': 0.0, 'saldo': saldo
                })

                # ========== PÓS-ENTREGA ==========
                tracker_pos = PaymentTracker(dia_pagamento, taxa_pos)
                tracker_pos.last_date = data_entrega
                prev_date = data_entrega
                cursor = data_entrega
                parcelas = 1

                while saldo > 0 and parcelas <= 420:
                    d_evt = adjust_day(cursor + relativedelta(months=1), dia_pagamento)

                    # (a) não-associados entre prev_date e d_evt — seguem iguais
                    for ev_nr in [e for e in post_nr if not e['assoc'] and prev_date < e['data'] < d_evt]:
                        juros, dias_corr, taxa_eff = tracker_pos.calculate(ev_nr['data'], saldo)
                        ipca_nr = saldo * TAXA_IPCA
                        extras_nr = [saldo * t['pct'] if t['periodo'] in ['pós-entrega da chave', 'ambos'] else 0.0 for t in taxas_extras]
                        total_taxas_nr = sum(extras_nr) + ipca_nr
                        abat_nr = ev_nr['valor'] - juros - total_taxas_nr
                        saldo -= abat_nr
                        eventos.append({
                            'data': ev_nr['data'], 'parcela': '', 'tipo': ev_nr['tipo'], 'valor': ev_nr['valor'],
                            'juros': juros, 'dias_corridos': dias_corr, 'taxa_efetiva': taxa_eff,
                            'incc': 0.0, 'ipca': ipca_nr, 'taxas_extra': extras_nr,
                            'Total de mudança (R$)': -abat_nr, 'saldo': saldo
                        })

                    # >>> ORDEM CORRETA A PARTIR DAQUI <<<

                    # (b) parcela mensal pós — calcula encargos ANTES dos associados
                    juros, dias_corr, taxa_eff = tracker_pos.calculate(d_evt, saldo)
                    ipca = saldo * TAXA_IPCA
                    extras = [saldo * t['pct'] if t['periodo'] in ['pós-entrega da chave', 'ambos'] else 0.0 for t in taxas_extras]
                    total_taxas = sum(extras) + ipca
                    valor_total = capacidade_pos
                    abat_mes = valor_total - juros - total_taxas
                    saldo -= abat_mes
                    eventos.append({
                        'data': d_evt, 'parcela': parcelas, 'tipo': 'Pós-Entrega', 'valor': valor_total,
                        'juros': juros, 'dias_corridos': dias_corr, 'taxa_efetiva': taxa_eff,
                        'incc': 0.0, 'ipca': ipca, 'taxas_extra': extras,
                        'Total de mudança (R$)': -abat_mes, 'saldo': saldo
                    })
                    parcelas += 1

                    # (c) associados do dia — linhas separadas, zerando encargos, DEPOIS da parcela
                    associados = [e for e in post_nr if e['assoc'] and e['data'] == d_evt]
                    for ev_as in associados:
                        abat_assoc = ev_as['valor']
                        saldo -= abat_assoc
                        eventos.append({
                            'data': d_evt, 'parcela': '', 'tipo': ev_as['tipo'] + " (Associado)", 'valor': ev_as['valor'],
                            'juros': 0.0, 'dias_corridos': 0, 'taxa_efetiva': 0.0,
                            'incc': 0.0, 'ipca': 0.0, 'taxas_extra': [0.0]*len(taxas_extras),
                            'Total de mudança (R$)': -abat_assoc, 'saldo': saldo
                        })

                    prev_date = d_evt
                    cursor = d_evt


                # --- Rotulagem k/N baseada nos eventos EFETIVOS (pré+pós, associados e não) ---
                def _is_semi(t):   return str(t).startswith("Pagamento Semestral")
                def _is_annual(t): return str(t).startswith("Pagamento Anual")

                eventos_sorted = sorted(eventos, key=lambda x: x['data'])

                semi_idxs   = [i for i,ev in enumerate(eventos_sorted) if _is_semi(ev['tipo'])]
                annual_idxs = [i for i,ev in enumerate(eventos_sorted) if _is_annual(ev['tipo'])]

                semi_N   = len(semi_idxs)
                annual_N = len(annual_idxs)

                # zera 'parcela' textual para não conflitar com parcelas mensais (inteiros)
                for ev in eventos_sorted:
                    if not isinstance(ev.get('parcela',''), int):
                        ev['parcela'] = ''

                for k, i in enumerate(semi_idxs, start=1):
                    eventos_sorted[i]['parcela'] = f"{k}/{semi_N}" if semi_N else ''
                for k, i in enumerate(annual_idxs, start=1):
                    eventos_sorted[i]['parcela'] = f"{k}/{annual_N}" if annual_N else ''

                # ========== Excel ==========
                wb = Workbook()
                ws = wb.active
                ws.title = f"Financ-{cliente}"[:31]
                headers = ["Data","Parcela","Tipo","Dias no Mês","Dias Corridos","Taxa Efetiva","Valor Pago (R$)",
                           "Juros (R$)","INCC (R$)","IPCA (R$)"]
                headers += [f"Taxa {i+1} (R$)" for i in range(len(taxas_extras))]
                headers += ["Total de adições e subtrações (R$)","Saldo Devedor (R$)"]

                # Cabeçalho
                for i, h in enumerate(headers, 1):
                    cell = ws.cell(row=1, column=i, value=h)
                    cell.fill = HEADER_FILL
                    cell.font = Font(bold=True)

                # Linha inicial (saldo)
                ws.append(["-"]*(len(headers)-1) + [valor_imovel])

                # Eventos ordenados
                for ev in eventos_sorted:
                    row = [
                        ev['data'],
                        ev.get('parcela', ''),
                        ev['tipo'],
                        days_in_month(ev['data']),
                        ev.get('dias_corridos', ''),
                        ev.get('taxa_efetiva', ''),
                        ev.get('valor', 0),
                        ev.get('juros', 0),
                        ev.get('incc', 0),
                        ev.get('ipca', 0)
                    ]
                    row += ev.get('taxas_extra', []) + [ev.get('Total de mudança (R$)', 0), ev.get('saldo', 0)]
                    ws.append(row)

                # Ajuste largura
                for col_cells in ws.columns:
                    max_length = 0
                    column = get_column_letter(col_cells[0].column)
                    for cell in col_cells:
                        if cell.value is not None:
                            max_length = max(max_length, len(str(cell.value)))
                    ws.column_dimensions[column].width = max_length + 2

                # Formatos
                for col_idx, h in enumerate(headers, start=1):
                    for row_idx in range(2, ws.max_row + 1):
                        cell = ws.cell(row=row_idx, column=col_idx)
                        if h == "Data":
                            cell.number_format = DATE_FORMAT
                        elif h in ["Parcela", "Dias no Mês", "Dias Corridos"]:
                            cell.number_format = '0'
                        elif h == "Taxa Efetiva":
                            cell.number_format = PERCENT_FORMAT
                        else:
                            cell.number_format = CURRENCY_FORMAT

                # Coloração por sinal na coluna "Total de adições e subtrações (R$)"
                total_col_idx = headers.index("Total de adições e subtrações (R$)") + 1
                for row_idx in range(2, ws.max_row + 1):
                    cell = ws.cell(row=row_idx, column=total_col_idx)
                    try:
                        val = float(cell.value)
                        if val < 0:  # abatimento => verde
                            cell.font = Font(color=GREEN_COLOR)
                        elif val > 0:  # adição => vermelho
                            cell.font = Font(color=RED_COLOR)
                    except (TypeError, ValueError):
                        pass

                # Aviso de limite
                if parcelas >= 420 and saldo > 0:
                    st.error(
                        f"Financiamento de {cliente} não é possível: excede 420 parcelas e ainda sobra saldo. "
                        f"Restante: R${saldo:.2f}."
                    )

                # Download
                buf = BytesIO()
                wb.save(buf)
                buf.seek(0)
                st.success("Simulação concluída! Baixe o Excel abaixo.")
                st.download_button("Download Excel",
                                   data=buf,
                                   file_name=f"Financiamento {cliente or 'Cliente'}.xlsx",
                                   mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            except Exception as e:
                st.error("Ocorreu um erro ao gerar a planilha.")
                st.exception(e)

# ==========================
# Main
# ==========================
def main():
    if 'authenticated' not in st.session_state:
        st.session_state.authenticated = False  

    # Logo global (sempre renderiza). Protegido para não quebrar o app.
    try:
        add_corner_image(LOGO_PATH, width_px=130, corner="top-left", offset_x=80, offset_y=70)
    except Exception as _:
        pass

    if not st.session_state.authenticated:
        login_screen()
        st.stop()  # não renderiza o resto enquanto não logar
    app_body()

if __name__ == "__main__":
    main()
